#!/usr/bin/env python3
"""Import personnel users from the monthly phone statistics workbook."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import select

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

load_dotenv(ROOT_DIR / ".env")

from app.database_models import User  # noqa: E402
from app.db import get_session_factory, init_db  # noqa: E402
from app.services.auth_service import hash_password  # noqa: E402

DEFAULT_WORKBOOK = (
    "/Users/xx/Library/Containers/com.tencent.WeWorkMac/Data/Documents/Profiles/"
    "4F0E838F05CBB3B62FD6C974A6A63862/Caches/Files/2026-05/"
    "e7bd728746c682aefb2864b6da1e309c/202604号码统计.xlsx"
)


def normalize_phone(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace(" ", "").replace("-", "")


def read_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    required = ["姓名", "手机号", "一级部门", "二级部门"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"Excel 缺少列：{', '.join(missing)}")

    index = {name: headers.index(name) for name in required}
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[index["姓名"]] or "").strip()
        phone = normalize_phone(row[index["手机号"]])
        dept1 = str(row[index["一级部门"]] or "").strip()
        dept2 = str(row[index["二级部门"]] or "").strip()
        if not name and not phone:
            continue
        if len(phone) != 11 or not phone.isdigit():
            raise RuntimeError(f"第 {row_num} 行手机号不正确：{phone or '<空>'}")
        if phone in seen:
            raise RuntimeError(f"Excel 内手机号重复：{phone}")
        seen.add(phone)
        rows.append(
            {
                "name": name or phone,
                "phone": phone,
                "department_level1": dept1,
                "department_level2": dept2,
            }
        )
    return rows


def import_users(workbook_path: Path, *, reset_passwords: bool = False) -> dict[str, int]:
    rows = read_rows(workbook_path)
    created = 0
    updated = 0
    skipped_passwords = 0

    init_db()
    with get_session_factory()() as db:
        for row in rows:
            phone = row["phone"]
            user = db.scalar(select(User).where(User.phone == phone))
            if user is None:
                user = db.scalar(select(User).where(User.username == phone))

            if user is None:
                user = User(
                    username=phone,
                    phone=phone,
                    password_hash=hash_password(phone[-6:]),
                    display_name=row["name"],
                    department_level1=row["department_level1"] or None,
                    department_level2=row["department_level2"] or None,
                    source="local",
                    status="active",
                    must_change_password=True,
                    is_active=True,
                )
                db.add(user)
                created += 1
                continue

            user.username = phone
            user.phone = phone
            user.display_name = row["name"]
            user.department_level1 = row["department_level1"] or None
            user.department_level2 = row["department_level2"] or None
            user.source = user.source or "local"
            user.status = "active"
            user.is_active = True
            if reset_passwords:
                user.password_hash = hash_password(phone[-6:])
                user.must_change_password = True
            else:
                skipped_passwords += 1
            db.add(user)
            updated += 1

        db.commit()

    return {
        "total": len(rows),
        "created": created,
        "updated": updated,
        "skipped_passwords": skipped_passwords,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import personnel login users from an Excel workbook.")
    parser.add_argument("workbook", nargs="?", default=DEFAULT_WORKBOOK, help="Path to 202604号码统计.xlsx")
    parser.add_argument(
        "--reset-passwords",
        action="store_true",
        help="Reset existing users to phone last 6 digits.",
    )
    args = parser.parse_args()

    result = import_users(Path(args.workbook), reset_passwords=args.reset_passwords)
    print(
        "导入完成："
        f"总数 {result['total']}，"
        f"新增 {result['created']}，"
        f"更新 {result['updated']}，"
        f"未重置密码 {result['skipped_passwords']}"
    )


if __name__ == "__main__":
    main()
